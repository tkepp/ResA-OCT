#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import glob
import numpy as np
import nibabel as nib
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers, Font

# turn off all warnings
import warnings
warnings.filterwarnings("ignore")

from utils import remove_subjects
from evaluation import calculate_segm_metrics, calculate_regr_metrics

def remap_labels(seg):
    out_seg = np.zeros_like(seg)
    out_seg[seg == 1] = 1 # ILM <-> RNFL-GCL (RNFL)
    out_seg[seg == 2] = 2 # RNFL-GCL <-> GCL-IPL (GCL)
    out_seg[seg == 3] = 2 # GCL-IPL <-> IPL-INL (IPL)
    # ==> RNFL-GCL <=> IPL-INL (GCL+IPL)
    out_seg[seg == 4] = 3 # IPL-INL <-> INL-OPL (INL)
    out_seg[seg == 5] = 4 # INL-OPL <-> OPL-HFL (OPL)
    out_seg[seg == 6] = 5 # OPL-HFL <-> BMEIS (ONL)
    out_seg[seg == 7] = 6 # BMEIS <-> IS/OSJ (IS/OS)
    out_seg[seg == 8] = 6 # IS/OSJ <-> IB_OPR (Outer segment)
    # ==> BMEIS <=> IB_OPR (IS/OS + Outer segment)
    out_seg[seg == 9] = 7 # IB_OPR <-> OB_OPR (OPR)
    out_seg[seg == 10] = 7 # OB_OPR <-> IB_RPE (Subretinal virtual space)
    # ==> IB_OPR <=> IB_RPE (OPR+ Subretinal virtual space)
    out_seg[seg == 11] = 8 # IB_RPE <-> OB_RPE (RPE)
    return out_seg


gt_path = ""


# folders of final paper submission
result_dirs = ['','']

# iterate over all result dirs
for idx1, result_dir in enumerate(result_dirs):

    # select inr or linear/reg dir by index
    if idx1 <= 3:
        result_path = os.path.join("", result_dir)
    else:
        result_path = os.path.join("", result_dir)
    idx_known = np.arange(0, 31, 2)
    idx_interp = np.arange(1, 31, 2)

    mses = []
    maes = []
    psnrs = []
    ssims = []
    lpipss = []
    dices = []
    assds = []
    hds95 = []
    hds = []
    subject_names = []

    #load data
    if 'registration' == result_dir:
        recon_paths = sorted(glob.glob(os.path.join(result_path, '**', 'oct_flat*registration.nii.gz'), recursive=True))
        seg_paths = sorted(glob.glob(os.path.join(result_path, '**', 'oct_seg_flat*registration.nii.gz'), recursive=True))
    elif 'linear' == result_dir:
        recon_paths = sorted(glob.glob(os.path.join(result_path, '**', 'oct_flat*.nii.gz'), recursive=True))
        seg_paths = sorted(glob.glob(os.path.join(result_path, '**', 'oct_seg_flat*.nii.gz'), recursive=True))
    else:
        recon_paths = sorted(glob.glob(os.path.join(result_path, '**', '*vol.nii.gz'), recursive=True))
        seg_paths = sorted(glob.glob(os.path.join(result_path, '**', '*seg.nii.gz'), recursive=True))

    # ignore subject 004
    ignored_subjects = ['sub004']
    gt_paths = remove_subjects(sorted(glob.glob(os.path.join(gt_path,'**','oct_flat_subsmpl16_o*.nii.gz'), recursive=True)), ignored_subjects=ignored_subjects)
    gt_seg_paths = remove_subjects(sorted(glob.glob(os.path.join(gt_path,'**','oct_seg_flat_subsmpl16_o*.nii.gz'), recursive=True)), ignored_subjects=ignored_subjects)


    for idx2 in tqdm(range(len(gt_paths))):

        # computation of regr and seg metrics
        nifti = nib.load(gt_paths[idx2])
        sx, sy, sz = nifti.header.get_zooms()
        gt_vol = np.asarray(nifti.get_fdata()) / 255.
        gt_seg_vol = remap_labels(np.asarray(nib.load(gt_seg_paths[idx2]).get_fdata()))
        pred_vol = np.asarray(nib.load(recon_paths[idx2]).get_fdata())
        if idx1 > 3:
            pred_vol /= 255.
        pred_seg_vol = remap_labels(np.asarray(nib.load(seg_paths[idx2]).get_fdata()))

        subject_name = f'{gt_paths[idx2].split("/")[-3].capitalize()}_{gt_paths[idx2].split("/")[-2]}'

        metrics_recon = calculate_regr_metrics(gt_vol[..., idx_interp], pred_vol[..., idx_interp])
        mses.append(metrics_recon['MSE'] * 100)
        maes.append(metrics_recon['MAE'] * 100)
        psnrs.append(metrics_recon['PSNR'])
        ssims.append(metrics_recon['SSIM'] * 100)
        lpipss.append(metrics_recon['LPIPS'])


        metrics_seg = calculate_segm_metrics(gt_seg_vol[..., idx_interp],
                                                  pred_seg_vol[..., idx_interp],
                                                  sampling=np.array([sx, sy]))

        dices.append(metrics_seg['DICE'] * 100) # save in %
        assds.append(metrics_seg['ASSD'] * 1000) # save in µm
        hds95.append(metrics_seg['HD 95'] * 1000) # save in µm
        hds.append(metrics_seg['HD'] * 1000) # save in µm

        subject_names.append(subject_name)

    # Create DataFrame
    df = pd.DataFrame(
        {'Subject Name': subject_names, 'MSE': mses, 'MAE': maes, 'PSNR': psnrs, 'SSIM': ssims, 'LPIPS': lpipss,
         'Dice': dices, 'ASSD': assds, 'HD95': hds95, 'HD': hds})

    # Compute and add mean, std, and median rows for test and train sets
    summary_stats = {
        'Subject Name': ['Test Mean', 'Test Std', 'Test Median', 'Train Mean', 'Train Std', 'Train Median'],
        'MSE': [df['MSE'][:20].mean(), df['MSE'][:20].std(), df['MSE'][:20].median(), df['MSE'][20:].mean(),
                df['MSE'][20:].std(), df['MSE'][20:].median()],
        'MAE': [df['MAE'][:20].mean(), df['MAE'][:20].std(), df['MAE'][:20].median(), df['MAE'][20:].mean(),
                df['MAE'][20:].std(), df['MAE'][20:].median()],
        'PSNR': [df['PSNR'][:20].mean(), df['PSNR'][:20].std(), df['PSNR'][:20].median(), df['PSNR'][20:].mean(),
                 df['PSNR'][20:].std(), df['PSNR'][20:].median()],
        'SSIM': [df['SSIM'][:20].mean(), df['SSIM'][:20].std(), df['SSIM'][:20].median(), df['SSIM'][20:].mean(),
                 df['SSIM'][20:].std(), df['SSIM'][20:].median()],
        'LPIPS': [df['LPIPS'][:20].mean(), df['LPIPS'][:20].std(), df['LPIPS'][:20].median(),
                  df['LPIPS'][20:].mean(), df['LPIPS'][20:].std(), df['LPIPS'][20:].median()],
        'Dice': [df['Dice'][:20].mean(), df['Dice'][:20].std(), df['Dice'][:20].median(), df['Dice'][20:].mean(),
                 df['Dice'][20:].std(), df['Dice'][20:].median()],
        'ASSD': [df['ASSD'][:20].mean(), df['ASSD'][:20].std(), df['ASSD'][:20].median(), df['ASSD'][20:].mean(),
                 df['ASSD'][20:].std(), df['ASSD'][20:].median()],
        'HD95': [df['HD95'][:20].mean(), df['HD95'][:20].std(), df['HD95'][:20].median(), df['HD95'][20:].mean(),
                 df['HD95'][20:].std(), df['HD95'][20:].median()],
        'HD': [df['HD'][:20].mean(), df['HD'][:20].std(), df['HD'][:20].median(), df['HD'][20:].mean(),
               df['HD'][20:].std(), df['HD'][20:].median()]
    }
    summary_df = pd.DataFrame(summary_stats)

    # Append summary statistics to the DataFrame
    df = pd.concat([df, summary_df], ignore_index=True)

    if idx1 <= 3:
        out_xls_path = os.path.join(result_path, "results_new.xlsx")
    else:
        out_xls_path = f"./results_{result_dir}.xlsx"

    # Save the DataFrame to an Excel file without formatting
    df.to_excel(out_xls_path, index=False)

    # Load the workbook and select the active sheet
    wb = load_workbook(out_xls_path)
    ws = wb.active

    # Define the number formats
    number_format_2_decimal = '0.00'
    number_format_1_decimal = '0.0'
    bold_font = Font(bold=True)

    # Apply the number format to the specific columns (MSE and MAE with three decimal places, the rest with two decimal places) and bold formatting to Mean and Std rows
    for row in ['Test Mean', 'Test Std', 'Test Median', 'Train Mean', 'Train Std', 'Train Median']:
        row_idx = df.index[df['Subject Name'] == row].tolist()[
                      0] + 2  # +2 because openpyxl is 1-based and there is a header row
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name in ['PSNR', 'ASSD', 'HD95', 'HD', 'Dice', 'SSIM']:
                ws.cell(row=row_idx, column=col_idx).number_format = number_format_1_decimal
            else:
                ws.cell(row=row_idx, column=col_idx).number_format = number_format_2_decimal
            if 'Mean' in row or 'Std' in row:
                ws.cell(row=row_idx, column=col_idx).font = bold_font

    # Save the formatted workbook
    wb.save(out_xls_path)